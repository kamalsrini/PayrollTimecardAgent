#!/usr/bin/env python3
"""
Extract employee names and total hours from timesheet Excel files
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass
from datetime import datetime

@dataclass
class EmployeeTimesheetData:
    """Data class for employee timesheet information"""
    file_name: str
    employee_name: str
    total_hours: float
    period: str
    title: str
    project_manager: str
    client: str
    firm: str
    extraction_date: str

class TimesheetExtractor:
    """Extract data from timesheet Excel files"""
    
    def __init__(self):
        self.timesheet_files = []
        self.extracted_data = []
    
    def find_timesheet_files(self) -> List[str]:
        """Find all timesheet Excel files in current directory"""
        timesheet_files = []
        
        for file in os.listdir("."):
            if file.endswith(('.xlsx', '.xls')) and 'Time Sheet' in file:
                timesheet_files.append(file)
        
        # Sort files for consistent processing
        timesheet_files.sort()
        self.timesheet_files = timesheet_files
        return timesheet_files
    
    def extract_employee_data_from_sheet(self, workbook, sheet_name: str) -> Optional[EmployeeTimesheetData]:
        """Extract employee data from a specific sheet"""
        try:
            sheet = workbook[sheet_name]
            
            # Initialize data
            employee_name = ""
            total_hours = 0.0
            period = ""
            title = ""
            project_manager = ""
            client = ""
            firm = ""
            
            # Extract data by scanning the sheet
            for row_num in range(1, sheet.max_row + 1):
                for col_num in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = cell.value
                    
                    if value is None:
                        continue
                    
                    value_str = str(value).strip()
                    
                    # Extract employee name
                    if value_str == "Name " and col_num < sheet.max_column:
                        name_cell = sheet.cell(row=row_num, column=col_num + 1)
                        if name_cell.value:
                            employee_name = str(name_cell.value).strip()
                    
                    # Also check for "Name" without trailing space
                    elif value_str == "Name" and col_num < sheet.max_column:
                        name_cell = sheet.cell(row=row_num, column=col_num + 1)
                        if name_cell.value:
                            employee_name = str(name_cell.value).strip()
                    
                    # Extract title
                    elif value_str == "Title " and col_num < sheet.max_column:
                        title_cell = sheet.cell(row=row_num, column=col_num + 1)
                        if title_cell.value:
                            title = str(title_cell.value).strip()
                    
                    # Also check for "Title" without trailing space
                    elif value_str == "Title" and col_num < sheet.max_column:
                        title_cell = sheet.cell(row=row_num, column=col_num + 1)
                        if title_cell.value:
                            title = str(title_cell.value).strip()
                    
                    # Extract period
                    elif value_str == "Period " and col_num < sheet.max_column:
                        period_cell = sheet.cell(row=row_num, column=col_num + 1)
                        if period_cell.value:
                            period = str(period_cell.value).strip()
                    
                    # Also check for "Period" without trailing space
                    elif value_str == "Period" and col_num < sheet.max_column:
                        period_cell = sheet.cell(row=row_num, column=col_num + 1)
                        if period_cell.value:
                            period = str(period_cell.value).strip()
                    
                    # Extract project manager
                    elif value_str == "Project Manager " and col_num < sheet.max_column:
                        pm_cell = sheet.cell(row=row_num, column=col_num + 1)
                        if pm_cell.value:
                            project_manager = str(pm_cell.value).strip()
                    
                    # Extract client
                    elif value_str == "Firm " and col_num < sheet.max_column:
                        firm_cell = sheet.cell(row=row_num, column=col_num + 1)
                        if firm_cell.value:
                            firm = str(firm_cell.value).strip()
                    
                    # Extract total hours
                    elif value_str == "Total Hours " and col_num < sheet.max_column:
                        # Check the next cell first
                        hours_cell = sheet.cell(row=row_num, column=col_num + 1)
                        if hours_cell.value:
                            try:
                                total_hours = float(hours_cell.value)
                            except (ValueError, TypeError):
                                total_hours = 0.0
                        else:
                            # If next cell is empty, check column 4 (where hours usually are)
                            hours_cell = sheet.cell(row=row_num, column=4)
                            if hours_cell.value:
                                try:
                                    total_hours = float(hours_cell.value)
                                except (ValueError, TypeError):
                                    total_hours = 0.0
                    
                    # Also check for "Total Hours" without trailing space
                    elif value_str == "Total Hours" and col_num < sheet.max_column:
                        # Check the next cell first
                        hours_cell = sheet.cell(row=row_num, column=col_num + 1)
                        if hours_cell.value:
                            try:
                                total_hours = float(hours_cell.value)
                            except (ValueError, TypeError):
                                total_hours = 0.0
                        else:
                            # If next cell is empty, check column 4 (where hours usually are)
                            hours_cell = sheet.cell(row=row_num, column=4)
                            if hours_cell.value:
                                try:
                                    total_hours = float(hours_cell.value)
                                except (ValueError, TypeError):
                                    total_hours = 0.0
            
            # Only return data if we found an employee name and hours
            if employee_name and total_hours > 0:
                return EmployeeTimesheetData(
                    file_name="",
                    employee_name=employee_name,
                    total_hours=total_hours,
                    period=period,
                    title=title,
                    project_manager=project_manager,
                    client=client,
                    firm=firm,
                    extraction_date=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                )
            
            return None
            
        except Exception as e:
            print(f"Error extracting data from sheet {sheet_name}: {e}")
            return None
    
    def extract_from_file(self, file_path: str) -> List[EmployeeTimesheetData]:
        """Extract data from a single timesheet file"""
        print(f"\nProcessing: {file_path}")
        
        try:
            workbook = load_workbook(file_path, data_only=True)
            extracted_employees = []
            
            # Process each sheet in the workbook
            for sheet_name in workbook.sheetnames:
                print(f"  Processing sheet: {sheet_name}")
                
                employee_data = self.extract_employee_data_from_sheet(workbook, sheet_name)
                
                if employee_data:
                    employee_data.file_name = file_path
                    extracted_employees.append(employee_data)
                    print(f"    ✓ Found: {employee_data.employee_name} - {employee_data.total_hours} hours")
                else:
                    print(f"    - No employee data found in sheet {sheet_name}")
            
            workbook.close()
            return extracted_employees
            
        except Exception as e:
            print(f"Error processing file {file_path}: {e}")
            return []
    
    def extract_all_timesheets(self) -> List[EmployeeTimesheetData]:
        """Extract data from all timesheet files"""
        print("=" * 80)
        print("TIMESHEET DATA EXTRACTION")
        print("=" * 80)
        
        # Find all timesheet files
        timesheet_files = self.find_timesheet_files()
        
        if not timesheet_files:
            print("No timesheet files found!")
            return []
        
        print(f"Found {len(timesheet_files)} timesheet files:")
        for file in timesheet_files:
            print(f"  - {file}")
        
        all_employees = []
        
        # Process each file
        for file_path in timesheet_files:
            employees = self.extract_from_file(file_path)
            all_employees.extend(employees)
        
        self.extracted_data = all_employees
        return all_employees
    
    def print_summary(self):
        """Print summary of extracted data"""
        if not self.extracted_data:
            print("No data extracted.")
            return
        
        print("\n" + "=" * 80)
        print("EXTRACTION SUMMARY")
        print("=" * 80)
        
        total_hours = 0
        unique_employees = set()
        
        for employee in self.extracted_data:
            print(f"Employee: {employee.employee_name}")
            print(f"  File: {employee.file_name}")
            print(f"  Hours: {employee.total_hours}")
            print(f"  Period: {employee.period}")
            print(f"  Title: {employee.title}")
            print(f"  Project Manager: {employee.project_manager}")
            print(f"  Client/Firm: {employee.client} / {employee.firm}")
            print()
            
            total_hours += employee.total_hours
            unique_employees.add(employee.employee_name)
        
        print(f"Total Employees: {len(unique_employees)}")
        print(f"Total Hours: {total_hours}")
        print(f"Extraction Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    def export_to_csv(self, filename: str = "extracted_timesheet_data.csv"):
        """Export extracted data to CSV"""
        if not self.extracted_data:
            print("No data to export.")
            return
        
        # Convert to DataFrame
        data = []
        for employee in self.extracted_data:
            data.append({
                'File Name': employee.file_name,
                'Employee Name': employee.employee_name,
                'Total Hours': employee.total_hours,
                'Period': employee.period,
                'Title': employee.title,
                'Project Manager': employee.project_manager,
                'Client': employee.client,
                'Firm': employee.firm,
                'Extraction Date': employee.extraction_date
            })
        
        df = pd.DataFrame(data)
        df.to_csv(filename, index=False)
        print(f"\nData exported to: {filename}")
    
    def get_consolidated_data(self) -> Dict[str, float]:
        """Get consolidated data for payroll processing"""
        consolidated = {}
        
        for employee in self.extracted_data:
            name = employee.employee_name
            if name in consolidated:
                consolidated[name] += employee.total_hours
            else:
                consolidated[name] = employee.total_hours
        
        return consolidated

def main():
    """Main function to run the extraction"""
    extractor = TimesheetExtractor()
    
    # Extract data from all timesheets
    employees = extractor.extract_all_timesheets()
    
    if employees:
        # Print summary
        extractor.print_summary()
        
        # Export to CSV
        extractor.export_to_csv()
        
        # Show consolidated data
        print("\n" + "=" * 80)
        print("CONSOLIDATED DATA FOR PAYROLL")
        print("=" * 80)
        consolidated = extractor.get_consolidated_data()
        for name, hours in consolidated.items():
            print(f"{name}: {hours} hours")
        
        print(f"\nTotal unique employees: {len(consolidated)}")
        print(f"Total hours across all employees: {sum(consolidated.values())}")
    else:
        print("No employee data found in any timesheet files.")

if __name__ == "__main__":
    main()
