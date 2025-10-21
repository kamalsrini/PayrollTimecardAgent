#!/usr/bin/env python3
"""
Update the Payroll Totals file by adding a new sheet with extracted timesheet data
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime
import shutil

def create_new_sheet_with_timesheet_data():
    """Create a new sheet in the Payroll Totals file with extracted timesheet data"""
    
    print("=" * 80)
    print("UPDATING PAYROLL TOTALS WITH TIMESHEET DATA")
    print("=" * 80)
    
    # Check if required files exist
    if not os.path.exists("3 - Payroll Totals.xls"):
        print("❌ Error: '3 - Payroll Totals.xls' not found")
        return False
    
    if not os.path.exists("payroll_summary.csv"):
        print("❌ Error: 'payroll_summary.csv' not found. Run create_payroll_summary.py first.")
        return False
    
    try:
        # Read the extracted timesheet data
        print("📊 Reading extracted timesheet data...")
        timesheet_df = pd.read_csv("payroll_summary.csv")
        print(f"✓ Found {len(timesheet_df)} employee records")
        
        # Create a backup of the original file
        backup_file = f"3 - Payroll Totals_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
        shutil.copy2("3 - Payroll Totals.xls", backup_file)
        print(f"✓ Created backup: {backup_file}")
        
        # Convert .xls to .xlsx for easier manipulation
        print("🔄 Converting .xls to .xlsx for processing...")
        xlsx_file = "3 - Payroll Totals_temp.xlsx"
        
        # Read the original .xls file
        original_df = pd.read_excel("3 - Payroll Totals.xls", sheet_name='Payroll Totals')
        sheet1_df = pd.read_excel("3 - Payroll Totals.xls", sheet_name='Sheet1')
        
        # Create new Excel file with openpyxl
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Add the original Payroll Totals sheet
        print("📝 Adding original Payroll Totals sheet...")
        payroll_sheet = workbook.create_sheet("Payroll Totals")
        
        # Write original data to the sheet
        for row_idx, row in original_df.iterrows():
            for col_idx, value in enumerate(row):
                if pd.notna(value):
                    payroll_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)
        
        # Add the original Sheet1
        print("📝 Adding original Sheet1...")
        sheet1 = workbook.create_sheet("Sheet1")
        
        for row_idx, row in sheet1_df.iterrows():
            for col_idx, value in enumerate(row):
                if pd.notna(value):
                    sheet1.cell(row=row_idx + 1, column=col_idx + 1, value=value)
        
        # Create new sheet with timesheet data
        print("📝 Creating new 'Timesheet Data' sheet...")
        timesheet_sheet = workbook.create_sheet("Timesheet Data")
        
        # Add headers
        headers = ["Employee Name", "Total Hours", "Period", "Title", "Source File", "Extraction Date"]
        for col_idx, header in enumerate(headers, 1):
            timesheet_sheet.cell(row=1, column=col_idx, value=header)
            # Make headers bold
            timesheet_sheet.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)
        
        # Add data rows
        for row_idx, (_, row) in enumerate(timesheet_df.iterrows(), 2):
            timesheet_sheet.cell(row=row_idx, column=1, value=row['Employee Name'])
            timesheet_sheet.cell(row=row_idx, column=2, value=row['Total Hours'])
            timesheet_sheet.cell(row=row_idx, column=3, value=row['Period'])
            timesheet_sheet.cell(row=row_idx, column=4, value=row['Title'])
            timesheet_sheet.cell(row=row_idx, column=5, value=row['Source File'])
            timesheet_sheet.cell(row=row_idx, column=6, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-adjust column widths
        for column in timesheet_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            timesheet_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the new Excel file
        print("💾 Saving updated Payroll Totals file...")
        workbook.save(xlsx_file)
        workbook.close()
        
        # Replace the original file
        os.replace(xlsx_file, "3 - Payroll Totals.xlsx")
        
        print("✅ Successfully updated Payroll Totals file!")
        print(f"📁 New file: 3 - Payroll Totals.xlsx")
        print(f"📁 Backup: {backup_file}")
        print(f"📊 Added {len(timesheet_df)} employee records to new 'Timesheet Data' sheet")
        
        # Print summary
        print("\n" + "=" * 80)
        print("TIMESHEET DATA ADDED TO PAYROLL TOTALS")
        print("=" * 80)
        for _, row in timesheet_df.iterrows():
            print(f"{row['Employee Name']}: {row['Total Hours']} hours")
        
        total_hours = timesheet_df['Total Hours'].sum()
        print(f"\nTotal Hours: {total_hours}")
        print(f"Total Employees: {len(timesheet_df)}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error updating Payroll Totals: {e}")
        return False

def create_simple_timesheet_sheet():
    """Create a simple timesheet sheet with just Name and Total Hours"""
    
    print("\n" + "=" * 80)
    print("CREATING SIMPLE TIMESHEET SHEET")
    print("=" * 80)
    
    try:
        # Read the extracted timesheet data
        timesheet_df = pd.read_csv("payroll_summary.csv")
        
        # Create a new workbook for the simple timesheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Timesheet Summary"
        
        # Add headers
        sheet.cell(row=1, column=1, value="Employee Name").font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=1, column=2, value="Total Hours").font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=1, column=3, value="Period").font = openpyxl.styles.Font(bold=True)
        
        # Add data
        for row_idx, (_, row) in enumerate(timesheet_df.iterrows(), 2):
            sheet.cell(row=row_idx, column=1, value=row['Employee Name'])
            sheet.cell(row=row_idx, column=2, value=row['Total Hours'])
            sheet.cell(row=row_idx, column=3, value=row['Period'])
        
        # Add total row
        total_row = len(timesheet_df) + 3
        sheet.cell(row=total_row, column=1, value="TOTAL").font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=total_row, column=2, value=timesheet_df['Total Hours'].sum()).font = openpyxl.styles.Font(bold=True)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the simple timesheet
        workbook.save("Timesheet_Summary.xlsx")
        workbook.close()
        
        print("✅ Created simple timesheet summary: Timesheet_Summary.xlsx")
        
        return True
        
    except Exception as e:
        print(f"❌ Error creating simple timesheet: {e}")
        return False

def main():
    """Main function"""
    print("🚀 Starting Payroll Totals update process...")
    
    # Create the updated Payroll Totals file
    success1 = create_new_sheet_with_timesheet_data()
    
    # Create a simple timesheet summary
    success2 = create_simple_timesheet_sheet()
    
    if success1 and success2:
        print("\n🎉 All operations completed successfully!")
        print("\nFiles created:")
        print("  📁 3 - Payroll Totals.xlsx (updated with new sheet)")
        print("  📁 Timesheet_Summary.xlsx (simple summary)")
        print("  📁 3 - Payroll Totals_backup_*.xls (backup of original)")
    else:
        print("\n❌ Some operations failed. Check the error messages above.")

if __name__ == "__main__":
    main()
